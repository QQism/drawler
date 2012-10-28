from openpyxl import load_workbook
from lexis.models import Category, Word, MeanWord, WordNode
import re

def open_sheet(file_path, sheet_name):
    print 'Opening file...'
    wb = load_workbook(file_path)
    print 'File loaded!'
    ws = wb.get_sheet_by_name(sheet_name)
    return ws

def insert_dict_into_database(ws):
    read = True
    line = 0
    while read:
        value = ws.cell(row=line, column=0).value
        if value is not None and value != '':
            word, created = Word.objects.get_or_create(plain=value)
            print ''.join([str(line), ' ', value.encode('utf-8')])

            word_categories = word.categories.all();

            for column in range(1,4):
                cat_value = ws.cell(row=line, column=column).value
                insert_categories(word, word_categories, cat_value)

            word.save()
            line += 1
        else:
            read = False

def insert_categories(word, word_categories, value):
    if value is not None and value.strip() != '':
        cat_keys = [cat_key.strip().lower() for cat_key in value.split(',')]
        for cat_key in cat_keys:
            cat, created = Category.objects.get_or_create(code=cat_key)
            if cat not in word_categories:
                mean_word, created = MeanWord.objects.get_or_create(word=word, category=cat)


def import_dictionary(file_path=''):
    if file_path == '':
        print 'Dictionary path: '
        file_path = raw_input()
    print 'Sheet: (Sheet2)'
    sheet_name = raw_input()

    if sheet_name == '':
        sheet_name = 'Sheet2'
    ws = open_sheet(file_path, sheet_name)

    insert_dict_into_database(ws)
    print 'Import successfully'

def insert_phrases_into_database(ws):
    read = True
    # 1654 Ho*n, bad data, adj
    # 6809, bad data, verb
    # 9525, bad data, verb
    line = 9525
    get_node = lambda node_id: WordNode.objects.get(pk=node_id)

    while read:
        phrase = ws.cell(row=line, column=0).value
        if phrase is not None and phrase != '':
            phrase_cat = ws.cell(row=line, column=6).value.lower()
            root_word, created = Word.objects.get_or_create(plain=phrase_cat)
            root_cat, created = Category.objects.get_or_create(code=phrase_cat)
            root_mean_word, created = MeanWord.objects.get_or_create(word=root_word, \
                                                           category=root_cat)
            phrase_root_node = WordNode.add_root(word=root_mean_word)
            phrase_root_node = get_node(phrase_root_node.id)

            word_parts = [w.replace(']', '').replace('[', ' ').split(' ') \
                          for w in phrase.split(' ')]
            for word_part in word_parts:
                word = word_part[0].lower()
                cat_key = word_part[1].lower()
                word, created = Word.objects.get_or_create(plain=word)
                cat, created = Category.objects.get_or_create(code=cat_key)
                mean_word, created = MeanWord.objects.get_or_create(word=word,\
                                                                    category=cat)
                phrase_root_node.add_child(word=mean_word)

            phrase_root_node = get_node(phrase_root_node.id)
            print phrase_root_node.get_text().encode('utf-8')

            line += 1
        else:
            read = False

def import_phrases(file_path=''):
    if file_path == '':
        print 'Phrase path: '
        file_path = raw_input()
    print 'Sheet: (Sheet2)'
    #sheet_name = raw_input()

#    if sheet_name == '':
    sheet_name = 'Sheet2'
    ws = open_sheet(file_path, sheet_name)

    insert_phrases_into_database(ws)
    print 'Import successfully'
